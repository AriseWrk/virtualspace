import os
import uuid
from datetime import datetime
from flask import (Blueprint, render_template, redirect,
                   url_for, flash, request, abort, send_from_directory)
from flask_login import login_required, current_user
from app.extensions import db
from app.models.order import Order, OrderItem
from app.models.project import (Project, ChecklistItem, CableJournal,
                                 IPTable, ProjectPhoto, ProjectNote,
                                 ProjectDocument, ProjectOrder)
from app.models.warehouse import Item, StockMovement
from app.models.user import User
from functools import wraps
from app.models.service_task import ServiceTask, ServiceTaskEngineer
from sqlalchemy import func

engineer_bp = Blueprint("engineer", __name__, url_prefix="/engineer")

UPLOAD_FOLDER     = os.path.join("app", "static", "uploads", "projects")
ALLOWED_DOC_EXTS  = {"pdf", "dwg", "dxf", "doc", "docx", "xls", "xlsx",
                     "png", "jpg", "jpeg", "zip", "rar", "7z"}


# ─────────────────────────────────────────────────────────────────────────────
# Декораторы доступа
# ─────────────────────────────────────────────────────────────────────────────

def engineer_required(f):
    """Только инженер и директор."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if current_user.role not in ("engineer", "director"):
            flash("Доступ запрещён.", "error")
            return redirect(url_for("dashboard.index"))
        return f(*args, **kwargs)
    return login_required(decorated)


def project_access(f):
    """Любой авторизованный пользователь может зайти в проект."""
    @wraps(f)
    def decorated(*args, **kwargs):
        return f(*args, **kwargs)
    return login_required(decorated)


def _can_edit_engineer(project):
    """Инженер может редактировать только свой проект."""
    return current_user.role == "director" or project.engineer_id == current_user.id


def _can_edit_designer(project):
    """Проектировщик может редактировать только свой проект."""
    return current_user.role in ("director",) or \
           project.designer_id == current_user.id or \
           current_user.role == "designer"


def _can_edit_office():
    """Офис и директор могут создавать заказы."""
    return current_user.role in ("office", "director", "pts")


# ─────────────────────────────────────────────────────────────────────────────
# ДАШБОРД
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/")
@engineer_required
def dashboard():
    uid = current_user.id
    is_director = current_user.role == "director"

    # ── Проекты ──────────────────────────────────────────────────────────────
    if is_director:
        my_projects = Project.query.order_by(Project.updated_at.desc()).limit(6).all()
        open_orders = Order.query.filter(Order.status.in_(["new", "in_progress"])).count()
        total_orders = Order.query.count()
    else:
        my_projects = Project.query.filter_by(
            engineer_id=uid
        ).order_by(Project.updated_at.desc()).limit(6).all()
        open_orders = Order.query.filter(
            Order.assigned_to_id == uid,
            Order.status.in_(["new", "in_progress"])
        ).count()
        total_orders = Order.query.filter_by(assigned_to_id=uid).count()

    active_projects = sum(
        1 for p in my_projects if p.status in ("in_progress", "pnr")
    )

    # ── Склад — заказы назначенные мне ───────────────────────────────────────
    my_warehouse_orders = Order.query.filter(
        Order.assigned_to_id == uid,
        Order.status.in_(["new", "in_progress"])
    ).order_by(Order.created_at.desc()).limit(5).all() if not is_director else \
        Order.query.filter(Order.status.in_(["new", "in_progress"])).order_by(
            Order.created_at.desc()
        ).limit(5).all()

    # ── Склад — мои остатки (Items которые я чаще заказываю) ─────────────────
    # Берём top-10 позиций из моих выданных заказов
    from app.models.order import OrderItem
    from app.models.warehouse import Item as WItem
    my_item_ids = db.session.query(
        OrderItem.item_id,
        db.func.count(OrderItem.id).label("cnt")
    ).join(Order).filter(
        Order.assigned_to_id == uid
    ).group_by(OrderItem.item_id).order_by(
        db.func.count(OrderItem.id).desc()
    ).limit(10).all()

    my_items = []
    for row in my_item_ids:
        item = WItem.query.get(row.item_id)
        if item:
            my_items.append(item)

    # ── Резерв — сколько зарезервировано под мои заказы ──────────────────────
    my_reserved_items = WItem.query.filter(
        WItem.reserved_qty > 0
    ).order_by(WItem.name).limit(8).all() if not is_director else []

    # ── План-задания ──────────────────────────────────────────────────────────
    from app.models.service_task import ServiceTask, ServiceTaskEngineer
    task_ids_q = db.session.query(ServiceTaskEngineer.task_id).filter_by(
        engineer_id=uid
    ).subquery()
    my_tasks = ServiceTask.query.filter(
        ServiceTask.id.in_(task_ids_q),
        ServiceTask.status.in_(["new", "assigned", "in_progress"])
    ).order_by(
        ServiceTask.priority.desc(),
        ServiceTask.planned_date.asc()
    ).limit(8).all() if not is_director else \
        ServiceTask.query.filter(
            ServiceTask.status.in_(["new", "assigned", "in_progress"])
        ).order_by(ServiceTask.planned_date.asc()).limit(8).all()

    # ── Статистика ────────────────────────────────────────────────────────────
    stats = {
        "open_orders": open_orders,
        "total_orders": total_orders,
        "active_projects": active_projects,
        "my_tasks": len(my_tasks),
        "urgent_tasks": sum(1 for t in my_tasks if t.is_urgent),
    }

    return render_template(
        "engineer/dashboard.html",
        stats=stats,
        my_projects=my_projects,
        my_warehouse_orders=my_warehouse_orders,
        my_items=my_items,
        my_reserved_items=my_reserved_items,
        my_tasks=my_tasks,
    )


# ─────────────────────────────────────────────────────────────────────────────
# СПИСОК ПРОЕКТОВ — доступен всем ролям
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects")
@login_required
def projects():
    status_filter = request.args.get("status", "")
    q = request.args.get("q", "").strip()

    # Директор и офис видят все проекты
    # Инженер — свои, Проектировщик — свои, остальные — все
    role = current_user.role
    if role == "engineer":
        query = Project.query.filter_by(engineer_id=current_user.id)
    elif role == "designer":
        query = Project.query.filter_by(designer_id=current_user.id)
    else:
        query = Project.query

    if status_filter:
        query = query.filter_by(status=status_filter)
    if q:
        query = query.filter(
            db.or_(Project.name.ilike(f"%{q}%"), Project.client.ilike(f"%{q}%"))
        )

    all_projects = query.order_by(Project.updated_at.desc()).all()
    engineers = User.query.filter(User.role.in_(["engineer", "director"])).all()
    designers = User.query.filter(User.role.in_(["designer", "director"])).all()

    return render_template(
        "engineer/projects.html",
        projects=all_projects,
        status_filter=status_filter,
        q=q,
        engineers=engineers,
        designers=designers,
        status_choices=Project.STATUS_CHOICES,
    )


# ─────────────────────────────────────────────────────────────────────────────
# СОЗДАНИЕ ПРОЕКТА — директор, офис, инженер
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/create", methods=["GET", "POST"])
@login_required
def create_project():
    if current_user.role not in ("engineer", "director", "office", "pts"):
        flash("Доступ запрещён.", "error")
        return redirect(url_for("engineer.projects"))

    engineers = User.query.filter(User.role.in_(["engineer", "director"])).all()
    designers = User.query.filter(User.role.in_(["designer", "director"])).all()

    if request.method == "POST":
        project = Project(
            name=request.form["name"].strip(),
            address=request.form.get("address", "").strip() or None,
            client=request.form.get("client", "").strip() or None,
            description=request.form.get("description", "").strip() or None,
            status=request.form.get("status", "new"),
            engineer_id=request.form.get("engineer_id", type=int) or None,
            designer_id=request.form.get("designer_id", type=int) or None,
            created_by_id=current_user.id,
        )
        db.session.add(project)
        db.session.flush()

        # Стандартный чеклист инженера
        default_checklist = [
            ("Выезд на объект, обследование",       "preparation",   1),
            ("Согласование схемы размещения",        "preparation",   2),
            ("Получение материалов со склада",       "preparation",   3),
            ("Прокладка кабельных трасс",            "cabling",       4),
            ("Монтаж оборудования",                  "installation",  5),
            ("Подключение кабелей",                  "cabling",       6),
            ("Пуско-наладочные работы",              "commissioning", 7),
            ("Тестирование системы",                 "commissioning", 8),
            ("Оформление кабельного журнала",        "documentation", 9),
            ("Оформление исполнительной документации","documentation",10),
            ("Сдача объекта заказчику",              "handover",     11),
        ]
        for title, category, order in default_checklist:
            db.session.add(ChecklistItem(
                project_id=project.id,
                title=title, category=category, order=order,
            ))

        db.session.commit()
        flash(f"Проект «{project.name}» создан.", "success")
        return redirect(url_for("engineer.project_detail", project_id=project.id))

    return render_template("engineer/project_form.html",
                           project=None, engineers=engineers, designers=designers)


# ─────────────────────────────────────────────────────────────────────────────
# СТРАНИЦА ПРОЕКТА — все роли
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>")
@login_required
def project_detail(project_id):
    project = Project.query.get_or_404(project_id)

    # Заказы склада привязанные к этому проекту
    related_orders = [po.order for po in project.project_orders if po.order]

    # Все заказы склада для формы (офис)
    all_warehouse_items = Item.query.order_by(Item.name).all() if _can_edit_office() else []
    engineers_list = User.query.filter(User.role.in_(["engineer", "director"])).all()

    # Чеклист по категориям
    checklist_by_category = {}
    for item in project.checklist:
        cat = item.category_label
        checklist_by_category.setdefault(cat, []).append(item)

    # Документы по типам
    docs_by_type = {}
    for dt in ProjectDocument.DOC_TYPES:
        docs_by_type[dt] = project.docs_by_type(dt)

    return render_template(
        "engineer/project_detail.html",
        project=project,
        related_orders=related_orders,
        checklist_by_category=checklist_by_category,
        docs_by_type=docs_by_type,
        doc_types=ProjectDocument.DOC_TYPES,
        doc_icons=ProjectDocument.DOC_ICONS,
        all_warehouse_items=all_warehouse_items,
        engineers_list=engineers_list,
        can_edit_engineer=_can_edit_engineer(project),
        can_edit_designer=_can_edit_designer(project),
        can_edit_office=_can_edit_office(),
    )


@engineer_bp.route("/projects/<int:project_id>/edit", methods=["GET", "POST"])
@login_required
def edit_project(project_id):
    project = Project.query.get_or_404(project_id)
    if not _can_edit_engineer(project) and current_user.role not in ("office", "pts"):
        flash("Доступ запрещён.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    engineers = User.query.filter(User.role.in_(["engineer", "director"])).all()
    designers = User.query.filter(User.role.in_(["designer", "director"])).all()

    if request.method == "POST":
        project.name        = request.form["name"].strip()
        project.address     = request.form.get("address", "").strip() or None
        project.client      = request.form.get("client", "").strip() or None
        project.description = request.form.get("description", "").strip() or None
        project.status      = request.form.get("status", "new")
        project.engineer_id = request.form.get("engineer_id", type=int) or None
        project.designer_id = request.form.get("designer_id", type=int) or None
        project.updated_at  = datetime.utcnow()
        db.session.commit()
        flash("Проект обновлён.", "success")
        return redirect(url_for("engineer.project_detail", project_id=project.id))

    return render_template("engineer/project_form.html",
                           project=project, engineers=engineers, designers=designers)


# ─────────────────────────────────────────────────────────────────────────────
# СТРАНИЦА ЧЕКЛИСТА
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>/checklist")
@login_required
def project_checklist(project_id):
    project = Project.query.get_or_404(project_id)

    checklist_by_category = {}
    for item in project.checklist:
        cat = item.category_label
        checklist_by_category.setdefault(cat, []).append(item)

    return render_template(
        "engineer/project_checklist.html",
        project=project,
        checklist_by_category=checklist_by_category,
        can_edit_engineer=_can_edit_engineer(project),
    )

@engineer_bp.route("/projects/<int:project_id>/checklist/toggle/<int:item_id>", methods=["POST"])
@login_required
def toggle_checklist(project_id, item_id):
    project = Project.query.get_or_404(project_id)
    if not _can_edit_engineer(project):
        flash("Только инженер проекта может отмечать чеклист.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))
    item = ChecklistItem.query.get_or_404(item_id)
    item.is_done    = not item.is_done
    item.done_at    = datetime.utcnow() if item.is_done else None
    item.done_by_id = current_user.id  if item.is_done else None
    db.session.commit()
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#checklist")


@engineer_bp.route("/projects/<int:project_id>/checklist/add", methods=["POST"])
@login_required
def add_checklist_item(project_id):
    project = Project.query.get_or_404(project_id)
    if not _can_edit_engineer(project):
        flash("Доступ запрещён.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))
    title = request.form.get("title", "").strip()
    if title:
        max_order = db.session.query(
            db.func.max(ChecklistItem.order)
        ).filter_by(project_id=project_id).scalar() or 0
        db.session.add(ChecklistItem(
            project_id=project_id, title=title,
            category=request.form.get("category", "installation"),
            order=max_order + 1,
        ))
        db.session.commit()
        flash("Пункт добавлен.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#checklist")


# ─────────────────────────────────────────────────────────────────────────────
# КАБЕЛЬНЫЙ ЖУРНАЛ (инженер)
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>/cable/add", methods=["POST"])
@login_required
def add_cable(project_id):
    project = Project.query.get_or_404(project_id)
    if not _can_edit_engineer(project):
        flash("Доступ запрещён.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))
    db.session.add(CableJournal(
        project_id=project_id,
        number=request.form["number"].strip(),
        cable_type=request.form["cable_type"].strip(),
        from_point=request.form["from_point"].strip(),
        to_point=request.form["to_point"].strip(),
        length=float(request.form["length"]) if request.form.get("length") else None,
        section=request.form.get("section", "").strip() or None,
        status=request.form.get("status", "planned"),
        notes=request.form.get("notes", "").strip() or None,
    ))
    db.session.commit()
    flash("Кабель добавлен.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#cables")


@engineer_bp.route("/projects/<int:project_id>/cable/<int:cable_id>/status", methods=["POST"])
@login_required
def update_cable_status(project_id, cable_id):
    cable = CableJournal.query.get_or_404(cable_id)
    cable.status = request.form.get("status", cable.status)
    db.session.commit()
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#cables")


@engineer_bp.route("/projects/<int:project_id>/cable/<int:cable_id>/delete", methods=["POST"])
@login_required
def delete_cable(project_id, cable_id):
    db.session.delete(CableJournal.query.get_or_404(cable_id))
    db.session.commit()
    flash("Запись удалена.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#cables")


# ─────────────────────────────────────────────────────────────────────────────
# ТАБЛИЦА IP (инженер)
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>/ip/add", methods=["POST"])
@login_required
def add_ip(project_id):
    project = Project.query.get_or_404(project_id)
    if not _can_edit_engineer(project):
        flash("Доступ запрещён.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))
    db.session.add(IPTable(
        project_id=project_id,
        ip_address=request.form["ip_address"].strip(),
        mac_address=request.form.get("mac_address", "").strip() or None,
        device_name=request.form["device_name"].strip(),
        device_model=request.form.get("device_model", "").strip() or None,
        location=request.form.get("location", "").strip() or None,
        login=request.form.get("login", "").strip() or None,
        password=request.form.get("password", "").strip() or None,
        status=request.form.get("status", "active"),
        notes=request.form.get("notes", "").strip() or None,
    ))
    db.session.commit()
    flash("Устройство добавлено.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#iptable")


@engineer_bp.route("/projects/<int:project_id>/ip/<int:ip_id>/delete", methods=["POST"])
@login_required
def delete_ip(project_id, ip_id):
    db.session.delete(IPTable.query.get_or_404(ip_id))
    db.session.commit()
    flash("Запись удалена.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#iptable")


# ─────────────────────────────────────────────────────────────────────────────
# ФОТО (инженер)
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>/photo/upload", methods=["POST"])
@login_required
def upload_photo(project_id):
    project = Project.query.get_or_404(project_id)
    if not _can_edit_engineer(project):
        flash("Доступ запрещён.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    file = request.files.get("photo")
    if not file or file.filename == "":
        flash("Файл не выбран.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    ext = file.filename.rsplit(".", 1)[-1].lower()
    if ext not in {"jpg", "jpeg", "png", "gif", "webp"}:
        flash("Разрешены только изображения.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    folder = os.path.join(UPLOAD_FOLDER, str(project_id))
    os.makedirs(folder, exist_ok=True)
    filename = f"{uuid.uuid4().hex}.{ext}"
    file.save(os.path.join(folder, filename))

    db.session.add(ProjectPhoto(
        project_id=project_id, filename=filename,
        description=request.form.get("description", "").strip() or None,
        uploaded_by=current_user.id,
    ))
    db.session.commit()
    flash("Фото загружено.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#photos")


@engineer_bp.route("/projects/<int:project_id>/photo/<int:photo_id>/delete", methods=["POST"])
@login_required
def delete_photo(project_id, photo_id):
    photo = ProjectPhoto.query.get_or_404(photo_id)
    filepath = os.path.join(UPLOAD_FOLDER, str(project_id), photo.filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    db.session.delete(photo)
    db.session.commit()
    flash("Фото удалено.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#photos")


# ─────────────────────────────────────────────────────────────────────────────
# ЗАМЕТКИ (все)
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>/note/add", methods=["POST"])
@login_required
def add_note(project_id):
    text = request.form.get("text", "").strip()
    if text:
        db.session.add(ProjectNote(
            project_id=project_id, text=text,
            note_type=request.form.get("note_type", "note"),
            author_id=current_user.id,
        ))
        db.session.commit()
        flash("Заметка добавлена.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#notes")


# ─────────────────────────────────────────────────────────────────────────────
# ДОКУМЕНТАЦИЯ ПРОЕКТИРОВЩИКА
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>/docs/upload", methods=["POST"])
@login_required
def upload_document(project_id):
    if not _can_edit_designer(Project.query.get_or_404(project_id)):
        flash("Доступ запрещён. Только проектировщик проекта.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    file = request.files.get("file")
    if not file or file.filename == "":
        flash("Файл не выбран.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_DOC_EXTS:
        flash(f"Тип файла .{ext} не разрешён.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    doc_type = request.form.get("doc_type", "pdf")
    if doc_type not in ProjectDocument.DOC_TYPES:
        flash("Неверный тип документа.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    folder = os.path.join(UPLOAD_FOLDER, str(project_id), "docs")
    os.makedirs(folder, exist_ok=True)

    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(folder, filename)
    file.save(filepath)
    file_size = os.path.getsize(filepath)

    db.session.add(ProjectDocument(
        project_id=project_id,
        doc_type=doc_type,
        title=request.form.get("title", file.filename).strip(),
        filename=filename,
        original_name=file.filename,
        file_size=file_size,
        version=request.form.get("version", "").strip() or None,
        notes=request.form.get("notes", "").strip() or None,
        uploaded_by_id=current_user.id,
    ))
    db.session.commit()
    flash("Документ загружен.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#docs")


@engineer_bp.route("/projects/<int:project_id>/docs/<int:doc_id>/delete", methods=["POST"])
@login_required
def delete_document(project_id, doc_id):
    project = Project.query.get_or_404(project_id)
    if not _can_edit_designer(project):
        flash("Доступ запрещён.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    doc = ProjectDocument.query.get_or_404(doc_id)
    filepath = os.path.join(UPLOAD_FOLDER, str(project_id), "docs", doc.filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    db.session.delete(doc)
    db.session.commit()
    flash("Документ удалён.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#docs")


@engineer_bp.route("/projects/<int:project_id>/docs/<int:doc_id>/download")
@login_required
def download_document(project_id, doc_id):
    doc = ProjectDocument.query.get_or_404(doc_id)
    folder = os.path.join(UPLOAD_FOLDER, str(project_id), "docs")
    return send_from_directory(folder, doc.filename,
                               as_attachment=True, download_name=doc.original_name)


# ─────────────────────────────────────────────────────────────────────────────
# ЗАКАЗЫ СО СКЛАДА (офис)
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>/orders/create", methods=["POST"])
@login_required
def create_project_order(project_id):
    if not _can_edit_office():
        flash("Только офис может создавать заказы.", "error")
        return redirect(url_for("engineer.project_detail", project_id=project_id))

    project = Project.query.get_or_404(project_id)

    from sqlalchemy import func
    last = db.session.query(func.max(Order.number)).filter(
        Order.number.like("ЗКЗ-%")
    ).scalar()
    n = int(last.split("-")[-1]) + 1 if last else 1
    number = f"ЗКЗ-{n:04d}"

    # Назначаем инженера проекта как получателя
    order = Order(
        number=number,
        object_name=f"{project.name}" + (f" — {project.address}" if project.address else ""),
        notes=request.form.get("notes", "").strip() or None,
        status="new",
        created_by_id=current_user.id,
        assigned_to_id=project.engineer_id,
    )
    db.session.add(order)
    db.session.flush()

    item_ids  = request.form.getlist("item_id[]")
    quantities = request.form.getlist("quantity[]")

    for iid, qty_s in zip(item_ids, quantities):
        if not iid or not qty_s:
            continue
        qty = float(qty_s)
        db.session.add(OrderItem(order_id=order.id, item_id=int(iid), quantity=qty))
        item = Item.query.get(int(iid))
        if item:
            item.reserved_qty = (item.reserved_qty or 0) + qty

    # Привязываем заказ к проекту
    db.session.add(ProjectOrder(
        project_id=project_id,
        order_id=order.id,
        created_by_id=current_user.id,
        notes=f"Создан из проекта «{project.name}»",
    ))
    db.session.commit()
    flash(f"Заказ {number} создан и отправлен на склад.", "success")
    return redirect(url_for("engineer.project_detail", project_id=project_id) + "#orders")


# ─────────────────────────────────────────────────────────────────────────────
# ЭКСПОРТ ПРОЕКТА
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/projects/<int:project_id>/export")
@login_required
def export_project(project_id):
    import io
    import openpyxl
    from openpyxl.styles import Font
    from flask import send_file

    project = Project.query.get_or_404(project_id)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Объект"
    ws1.merge_cells("A1:D1")
    ws1["A1"] = f"Объект: {project.name}"
    ws1["A1"].font = Font(bold=True, size=14)
    ws1.append(["Адрес:",    project.address or "—"])
    ws1.append(["Заказчик:", project.client or "—"])
    ws1.append(["Статус:",   project.status_label])
    ws1.append(["Инженер:",  project.engineer.full_name if project.engineer else "—"])
    ws1.append(["Проектировщик:", project.designer.full_name if project.designer else "—"])
    ws1.append(["Прогресс:", f"{project.checklist_progress}%"])
    ws1.column_dimensions["A"].width = 25
    ws1.column_dimensions["B"].width = 40

    ws2 = wb.create_sheet("Чеклист")
    ws2.append(["#", "Пункт", "Категория", "Выполнен", "Дата", "Исполнитель"])
    for i, item in enumerate(project.checklist, 1):
        ws2.append([i, item.title, item.category_label,
                    "Да" if item.is_done else "Нет",
                    item.done_at.strftime("%d.%m.%Y") if item.done_at else "—",
                    item.done_by.full_name if item.done_by else "—"])

    ws3 = wb.create_sheet("Кабельный журнал")
    ws3.append(["№", "Тип", "От", "До", "Длина", "Сечение", "Статус"])
    for c in project.cable_journal:
        ws3.append([c.number, c.cable_type, c.from_point, c.to_point,
                    c.length or "—", c.section or "—", c.status_label])

    ws4 = wb.create_sheet("IP-таблица")
    ws4.append(["IP", "MAC", "Устройство", "Модель", "Расположение", "Статус"])
    for ip in project.ip_table:
        ws4.append([ip.ip_address, ip.mac_address or "—", ip.device_name,
                    ip.device_model or "—", ip.location or "—", ip.status_label])

    ws5 = wb.create_sheet("Документация")
    ws5.append(["Тип", "Название", "Файл", "Версия", "Загрузил", "Дата"])
    for doc in project.documents:
        ws5.append([doc.doc_type_label, doc.title, doc.original_name,
                    doc.version or "—",
                    doc.uploaded_by.full_name if doc.uploaded_by else "—",
                    doc.uploaded_at.strftime("%d.%m.%Y") if doc.uploaded_at else "—"])

    ws6 = wb.create_sheet("Заказы")
    ws6.append(["Номер", "Статус", "Позиций", "Дата"])
    for po in project.project_orders:
        o = po.order
        if o:
            ws6.append([o.number, o.status_label, o.total_items,
                        o.created_at.strftime("%d.%m.%Y")])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"project_{project.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ─────────────────────────────────────────────────────────────────────────────
# ОСТАЛЬНЫЕ РАЗДЕЛЫ
# ─────────────────────────────────────────────────────────────────────────────

@engineer_bp.route("/handbook")
@engineer_required
def handbook():
    return render_template("engineer/handbook.html")

@engineer_bp.route("/tests")
@engineer_required
def tests():
    return render_template("engineer/tests.html")

@engineer_bp.route("/regulations")
@engineer_required
def regulations():
    return render_template("engineer/regulations.html")

@engineer_bp.route("/software")
@engineer_required
def software():
    return render_template("engineer/software.html")

@engineer_bp.route("/equipment")
@engineer_required
def equipment():
    return render_template("engineer/equipment.html")