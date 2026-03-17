import os
import uuid
from datetime import datetime
from flask import (Blueprint, render_template, redirect,
                   url_for, flash, request, send_file, abort)
from flask_login import login_required, current_user
from app.extensions import db
from app.models.pts import (ObjectCategory, ServiceObject, ServiceRecord,
                             ObjectPassword, ObjectFile, ObjectEquipment)
from app.models.user import User
from functools import wraps
from app.models.order import Order

pts_bp = Blueprint("pts", __name__, url_prefix="/pts")
UPLOAD_FOLDER = os.path.join("app", "static", "uploads", "objects")


def pts_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if current_user.role not in ("office", "pts", "director", "engineer", "designer"):
            flash("Доступ запрещён.", "error")
            return redirect(url_for("dashboard.index"))
        return f(*args, **kwargs)
    return login_required(decorated)


# ===== ДАШБОРД =====

@pts_bp.route("/")
@pts_required
def dashboard():
    total_objects = ServiceObject.query.count()
    active_objects = ServiceObject.query.filter_by(status="active").count()

    upcoming_to = ServiceObject.query.filter(
        ServiceObject.next_to_date.isnot(None),
        ServiceObject.section == "service"
    ).order_by(ServiceObject.next_to_date.asc()).limit(5).all()

    recent_records = ServiceRecord.query.order_by(
        ServiceRecord.date.desc()
    ).limit(8).all()

    categories = ObjectCategory.query.order_by(ObjectCategory.order).all()

    stats = {
        "total": total_objects,
        "active": active_objects,
        "categories": len(categories),
        "installations": ServiceObject.query.filter_by(section="installation").count(),
    }
    return render_template("pts/dashboard.html",
                           stats=stats,
                           upcoming_to=upcoming_to,
                           recent_records=recent_records,
                           categories=categories,
                           now=datetime.utcnow())


# ===== ТЕХОБСЛУЖИВАНИЕ =====

@pts_bp.route("/service")
@pts_required
def service_list():
    status_filter = request.args.get("status", "")
    cat_filter = request.args.get("cat", type=int)

    query = ServiceObject.query.filter_by(section="service")
    if status_filter:
        query = query.filter_by(status=status_filter)
    if cat_filter:
        query = query.filter_by(category_id=cat_filter)

    objects = query.order_by(ServiceObject.name).all()
    categories = ObjectCategory.query.order_by(ObjectCategory.order).all()

    return render_template("pts/service_list.html",
                           objects=objects,
                           categories=categories,
                           status_filter=status_filter,
                           cat_filter=cat_filter,
                           now=datetime.utcnow())


# ===== МОНТАЖИ =====

@pts_bp.route("/installation")
@pts_required
def installation_list():
    status_filter = request.args.get("status", "")

    query = ServiceObject.query.filter_by(section="installation")
    if status_filter:
        query = query.filter_by(status=status_filter)

    objects = query.order_by(ServiceObject.updated_at.desc()).all()
    categories = ObjectCategory.query.order_by(ObjectCategory.order).all()

    return render_template("pts/installation_list.html",
                           objects=objects,
                           categories=categories,
                           status_filter=status_filter,
                           now=datetime.utcnow())


# ===== ВСЕ ОБЪЕКТЫ (старый роут, оставляем для совместимости) =====

@pts_bp.route("/objects")
@pts_required
def objects():
    categories = ObjectCategory.query.order_by(ObjectCategory.order).all()
    active_cat = request.args.get("cat", type=int)
    if not active_cat and categories:
        active_cat = categories[0].id

    current_category = None
    objects_list = []
    if active_cat:
        current_category = ObjectCategory.query.get(active_cat)
        if current_category:
            objects_list = current_category.objects.order_by(
                ServiceObject.name
            ).all()

    engineers = User.query.filter(User.role.in_(["engineer", "director"])).all()
    return render_template("pts/objects.html",
                           categories=categories,
                           active_cat=active_cat,
                           current_category=current_category,
                           objects_list=objects_list,
                           engineers=engineers)


# ===== КАТЕГОРИИ =====

@pts_bp.route("/objects/category/add", methods=["POST"])
@pts_required
def add_category():
    name = request.form.get("name", "").strip()
    if not name:
        flash("Название не может быть пустым.", "error")
        return redirect(url_for("pts.objects"))
    if ObjectCategory.query.filter_by(name=name).first():
        flash(f"Категория «{name}» уже существует.", "error")
        return redirect(url_for("pts.objects"))

    max_order = db.session.query(
        db.func.max(ObjectCategory.order)
    ).scalar() or 0
    cat = ObjectCategory(
        name=name,
        description=request.form.get("description", "").strip() or None,
        color=request.form.get("color", "#b98c50"),
        order=max_order + 1,
        created_by=current_user.id,
    )
    db.session.add(cat)
    db.session.commit()
    flash(f"Категория «{name}» добавлена.", "success")
    return redirect(url_for("pts.objects", cat=cat.id))


@pts_bp.route("/objects/category/<int:cat_id>/delete", methods=["POST"])
@pts_required
def delete_category(cat_id):
    cat = ObjectCategory.query.get_or_404(cat_id)
    if cat.objects.count() > 0:
        flash(f"Нельзя удалить — в категории {cat.objects.count()} объектов.", "error")
        return redirect(url_for("pts.objects", cat=cat_id))
    db.session.delete(cat)
    db.session.commit()
    flash(f"Категория «{cat.name}» удалена.", "success")
    return redirect(url_for("pts.objects"))


# ===== СОЗДАНИЕ ОБЪЕКТА =====

@pts_bp.route("/objects/create", methods=["GET", "POST"])
@pts_required
def create_object():
    categories = ObjectCategory.query.order_by(ObjectCategory.order).all()
    engineers = User.query.filter(User.role.in_(["engineer", "director"])).all()
    section = request.args.get("section", "service")

    if not categories:
        flash("Сначала создайте хотя бы одну категорию.", "warning")
        return redirect(url_for("pts.objects"))

    if request.method == "POST":
        section = request.form.get("section", "service")
        obj = ServiceObject(
            name=request.form["name"].strip(),
            address=request.form.get("address", "").strip() or None,
            category_id=int(request.form["category_id"]),
            section=section,
            client_name=request.form.get("client_name", "").strip() or None,
            client_contact=request.form.get("client_contact", "").strip() or None,
            client_phone=request.form.get("client_phone", "").strip() or None,
            client_email=request.form.get("client_email", "").strip() or None,
            systems=request.form.get("systems", "").strip() or None,
            engineer_id=request.form.get("engineer_id", type=int) or None,
            notes=request.form.get("notes", "").strip() or None,
            contract_number=request.form.get("contract_number", "").strip() or None,
        )

        if section == "installation":
            stage = request.form.get("installation_stage", "inst_survey")
            obj.installation_stage = stage
            obj.status = stage
            if request.form.get("estimate_sum"):
                obj.estimate_sum = float(request.form["estimate_sum"])
            if request.form.get("handover_date"):
                obj.handover_date = datetime.strptime(
                    request.form["handover_date"], "%Y-%m-%d")
        else:
            obj.status = request.form.get("status", "active")
            if request.form.get("next_to_date"):
                obj.next_to_date = datetime.strptime(
                    request.form["next_to_date"], "%Y-%m-%d")

        if request.form.get("commissioned_at"):
            obj.commissioned_at = datetime.strptime(
                request.form["commissioned_at"], "%Y-%m-%d")

        db.session.add(obj)
        db.session.commit()
        flash(f"Объект «{obj.name}» создан.", "success")
        return redirect(url_for("pts.object_detail", obj_id=obj.id))

    return render_template("pts/object_form.html",
                           obj=None,
                           categories=categories,
                           engineers=engineers,
                           section=section)


# ===== РЕДАКТИРОВАНИЕ ОБЪЕКТА =====

@pts_bp.route("/objects/<int:obj_id>/edit", methods=["GET", "POST"])
@pts_required
def edit_object(obj_id):
    obj = ServiceObject.query.get_or_404(obj_id)
    categories = ObjectCategory.query.order_by(ObjectCategory.order).all()
    engineers = User.query.filter(User.role.in_(["engineer", "director"])).all()

    if request.method == "POST":
        obj.name          = request.form["name"].strip()
        obj.address       = request.form.get("address", "").strip() or None
        obj.category_id   = int(request.form["category_id"])
        obj.client_name   = request.form.get("client_name", "").strip() or None
        obj.client_contact = request.form.get("client_contact", "").strip() or None
        obj.client_phone  = request.form.get("client_phone", "").strip() or None
        obj.client_email  = request.form.get("client_email", "").strip() or None
        obj.systems       = request.form.get("systems", "").strip() or None
        obj.engineer_id   = request.form.get("engineer_id", type=int) or None
        obj.notes         = request.form.get("notes", "").strip() or None
        obj.contract_number = request.form.get("contract_number", "").strip() or None
        obj.updated_at    = datetime.utcnow()

        if obj.section == "installation":
            stage = request.form.get("installation_stage", "inst_survey")
            obj.installation_stage = stage
            obj.status = stage
            if request.form.get("estimate_sum"):
                obj.estimate_sum = float(request.form["estimate_sum"])
            if request.form.get("handover_date"):
                obj.handover_date = datetime.strptime(
                    request.form["handover_date"], "%Y-%m-%d")
        else:
            obj.status = request.form.get("status", "active")
            if request.form.get("next_to_date"):
                obj.next_to_date = datetime.strptime(
                    request.form["next_to_date"], "%Y-%m-%d")

        if request.form.get("commissioned_at"):
            obj.commissioned_at = datetime.strptime(
                request.form["commissioned_at"], "%Y-%m-%d")

        db.session.commit()
        flash("Объект обновлён.", "success")
        return redirect(url_for("pts.object_detail", obj_id=obj.id))

    return render_template("pts/object_form.html",
                           obj=obj,
                           categories=categories,
                           engineers=engineers,
                           section=obj.section)


# ===== УДАЛЕНИЕ ОБЪЕКТА =====

@pts_bp.route("/objects/<int:obj_id>/delete", methods=["POST"])
@pts_required
def delete_object(obj_id):
    obj = ServiceObject.query.get_or_404(obj_id)
    section = obj.section
    name = obj.name

    # Удаляем файлы с диска
    folder = os.path.join(UPLOAD_FOLDER, str(obj_id))
    if os.path.exists(folder):
        import shutil
        shutil.rmtree(folder)

    db.session.delete(obj)
    db.session.commit()
    flash(f"Объект «{name}» удалён.", "success")

    if section == "installation":
        return redirect(url_for("pts.installation_list"))
    return redirect(url_for("pts.service_list"))


# ===== КАРТОЧКА ОБЪЕКТА =====

@pts_bp.route("/objects/<int:obj_id>")
@pts_required
def object_detail(obj_id):
    obj = ServiceObject.query.get_or_404(obj_id)

    # Заказы склада по этому объекту (ищем по названию объекта)
    obj_orders = Order.query.filter(
        Order.object_name.ilike(f"%{obj.name}%")
    ).order_by(Order.created_at.desc()).all()

    return render_template(
        "pts/object_detail.html",
        obj=obj,
        now=datetime.utcnow(),
        obj_orders=obj_orders,
    )
# ===== СЕРВИСНЫЕ ЗАПИСИ =====

@pts_bp.route("/objects/<int:obj_id>/record/add", methods=["POST"])
@pts_required
def add_record(obj_id):
    obj = ServiceObject.query.get_or_404(obj_id)
    date_str = request.form.get("date", "")
    record_date = datetime.strptime(date_str, "%Y-%m-%d") if date_str else datetime.utcnow()

    record = ServiceRecord(
        object_id=obj_id,
        record_type=request.form.get("record_type", "to"),
        status=request.form.get("status", "done"),
        date=record_date,
        description=request.form["description"].strip(),
        result=request.form.get("result", "").strip() or None,
        engineer_id=request.form.get("engineer_id", type=int) or None,
        created_by=current_user.id,
    )
    if request.form.get("next_to_date"):
        record.next_to_date = datetime.strptime(
            request.form["next_to_date"], "%Y-%m-%d")
        obj.next_to_date = record.next_to_date

    db.session.add(record)
    db.session.commit()
    flash("Запись добавлена.", "success")
    return redirect(url_for("pts.object_detail", obj_id=obj_id) + "#history")


@pts_bp.route("/objects/<int:obj_id>/record/<int:rec_id>/delete", methods=["POST"])
@pts_required
def delete_record(obj_id, rec_id):
    rec = ServiceRecord.query.get_or_404(rec_id)
    db.session.delete(rec)
    db.session.commit()
    flash("Запись удалена.", "success")
    return redirect(url_for("pts.object_detail", obj_id=obj_id) + "#history")


# ===== ПАРОЛИ =====

@pts_bp.route("/objects/<int:obj_id>/password/add", methods=["POST"])
@pts_required
def add_password(obj_id):
    pwd = ObjectPassword(
        object_id=obj_id,
        title=request.form["title"].strip(),
        login=request.form.get("login", "").strip() or None,
        password=request.form["password"].strip(),
        ip_or_url=request.form.get("ip_or_url", "").strip() or None,
        notes=request.form.get("notes", "").strip() or None,
    )
    db.session.add(pwd)
    db.session.commit()
    flash("Пароль добавлен.", "success")
    return redirect(url_for("pts.object_detail", obj_id=obj_id) + "#passwords")


@pts_bp.route("/objects/<int:obj_id>/password/<int:pwd_id>/delete", methods=["POST"])
@pts_required
def delete_password(obj_id, pwd_id):
    pwd = ObjectPassword.query.get_or_404(pwd_id)
    db.session.delete(pwd)
    db.session.commit()
    flash("Пароль удалён.", "success")
    return redirect(url_for("pts.object_detail", obj_id=obj_id) + "#passwords")


# ===== ФАЙЛЫ =====

@pts_bp.route("/objects/<int:obj_id>/file/upload", methods=["POST"])
@pts_required
def upload_file(obj_id):
    if "file" not in request.files or request.files["file"].filename == "":
        flash("Файл не выбран.", "error")
        return redirect(url_for("pts.object_detail", obj_id=obj_id))

    file = request.files["file"]
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else "bin"
    filename = f"{uuid.uuid4().hex}.{ext}"

    folder = os.path.join(UPLOAD_FOLDER, str(obj_id))
    os.makedirs(folder, exist_ok=True)
    file.save(os.path.join(folder, filename))

    obj_file = ObjectFile(
        object_id=obj_id,
        filename=filename,
        original_name=file.filename,
        file_type=request.form.get("file_type", "other"),
        description=request.form.get("description", "").strip() or None,
        version=request.form.get("version", "").strip() or None,
        uploaded_by=current_user.id,
    )
    db.session.add(obj_file)
    db.session.commit()
    flash(f"Файл «{file.filename}» загружен.", "success")
    return redirect(url_for("pts.object_detail", obj_id=obj_id) + "#files")


@pts_bp.route("/objects/<int:obj_id>/file/<int:file_id>/delete", methods=["POST"])
@pts_required
def delete_file(obj_id, file_id):
    obj_file = ObjectFile.query.get_or_404(file_id)
    filepath = os.path.join(UPLOAD_FOLDER, str(obj_id), obj_file.filename)
    if os.path.exists(filepath):
        os.remove(filepath)
    db.session.delete(obj_file)
    db.session.commit()
    flash("Файл удалён.", "success")
    return redirect(url_for("pts.object_detail", obj_id=obj_id) + "#files")


# ===== ОБОРУДОВАНИЕ =====

@pts_bp.route("/objects/<int:obj_id>/equipment/add", methods=["POST"])
@pts_required
def add_equipment(obj_id):
    eq = ObjectEquipment(
        object_id=obj_id,
        name=request.form["name"].strip(),
        model=request.form.get("model", "").strip() or None,
        serial_number=request.form.get("serial_number", "").strip() or None,
        quantity=int(request.form.get("quantity", 1)),
        location=request.form.get("location", "").strip() or None,
        notes=request.form.get("notes", "").strip() or None,
    )
    if request.form.get("installed_at"):
        eq.installed_at = datetime.strptime(request.form["installed_at"], "%Y-%m-%d")
    if request.form.get("warranty_until"):
        eq.warranty_until = datetime.strptime(request.form["warranty_until"], "%Y-%m-%d")

    db.session.add(eq)
    db.session.commit()
    flash("Оборудование добавлено.", "success")
    return redirect(url_for("pts.object_detail", obj_id=obj_id) + "#equipment")


@pts_bp.route("/objects/<int:obj_id>/equipment/<int:eq_id>/delete", methods=["POST"])
@pts_required
def delete_equipment(obj_id, eq_id):
    eq = ObjectEquipment.query.get_or_404(eq_id)
    db.session.delete(eq)
    db.session.commit()
    flash("Оборудование удалено.", "success")
    return redirect(url_for("pts.object_detail", obj_id=obj_id) + "#equipment")


# ===== ЭКСПОРТ =====

@pts_bp.route("/objects/<int:obj_id>/export")
@pts_required
def export_object(obj_id):
    import io
    import openpyxl
    from openpyxl.styles import Font

    obj = ServiceObject.query.get_or_404(obj_id)
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Объект"
    ws1.merge_cells("A1:C1")
    ws1["A1"] = f"Объект: {obj.name}"
    ws1["A1"].font = Font(bold=True, size=13)
    ws1.append(["Раздел:", "Техобслуживание" if obj.section == "service" else "Монтаж"])
    ws1.append(["Адрес:", obj.address or "—"])
    ws1.append(["Категория:", obj.category.name])
    ws1.append(["Статус:", obj.status_label])
    ws1.append(["Заказчик:", obj.client_name or "—"])
    ws1.append(["Контакт:", obj.client_contact or "—"])
    ws1.append(["Телефон:", obj.client_phone or "—"])
    ws1.append(["Инженер:", obj.engineer.full_name if obj.engineer else "—"])
    ws1.append(["Системы:", obj.systems or "—"])
    if obj.section == "installation":
        ws1.append(["Договор №:", obj.contract_number or "—"])
        ws1.append(["Смета:", f"{obj.estimate_sum:,.0f} ₽" if obj.estimate_sum else "—"])
        ws1.append(["Дата сдачи:", obj.handover_date.strftime("%d.%m.%Y") if obj.handover_date else "—"])
    if obj.commissioned_at:
        ws1.append(["Введён:", obj.commissioned_at.strftime("%d.%m.%Y")])
    if obj.next_to_date:
        ws1.append(["След. ТО:", obj.next_to_date.strftime("%d.%m.%Y")])
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 40

    ws2 = wb.create_sheet("История")
    ws2.append(["Дата", "Тип", "Статус", "Описание", "Результат", "Инженер"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for rec in obj.service_records:
        ws2.append([
            rec.date.strftime("%d.%m.%Y"), rec.type_label,
            rec.status_label, rec.description,
            rec.result or "—",
            rec.engineer.full_name if rec.engineer else "—",
        ])
    for col, w in zip("ABCDEF", [12, 15, 12, 50, 40, 25]):
        ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Оборудование")
    ws3.append(["Наименование", "Модель", "S/N", "Кол-во", "Расположение", "Установлен", "Гарантия до"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for eq in obj.equipment:
        ws3.append([
            eq.name, eq.model or "—", eq.serial_number or "—",
            eq.quantity, eq.location or "—",
            eq.installed_at.strftime("%d.%m.%Y") if eq.installed_at else "—",
            eq.warranty_until.strftime("%d.%m.%Y") if eq.warranty_until else "—",
        ])
    for col, w in zip("ABCDEFG", [30, 25, 20, 8, 25, 12, 12]):
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"object_{obj.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")