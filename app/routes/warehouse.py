from datetime import datetime, date
from io import BytesIO

from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, Response
from flask_login import login_required, current_user
from sqlalchemy import func
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.extensions import db
from app.models.warehouse import (
    Category, Item,
    StockMovement,
    Receipt, ReceiptItem,
    WriteOff, WriteOffItem,
    InventoryCheck, InventoryCheckItem,
)
from app.models.order import Order, OrderItem
from app.models.user import User

warehouse_bp = Blueprint("warehouse", __name__, url_prefix="/warehouse")
warehouse = warehouse_bp  # алиас для обратной совместимости

PER_PAGE = 50  # записей на страницу в журнале перемещений


# ─────────────────────────────────────────────────────────────────────────────
# Хелперы
# ─────────────────────────────────────────────────────────────────────────────

def _next_number(prefix: str, model, field="number") -> str:
    col = getattr(model, field)
    last = (
        db.session.query(func.max(col))
        .filter(col.like(f"{prefix}-%"))
        .scalar()
    )
    if last:
        try:
            n = int(last.split("-")[-1]) + 1
        except ValueError:
            n = 1
    else:
        n = 1
    return f"{prefix}-{n:04d}"


def _update_avg_cost(item: Item, new_qty: float, new_unit_cost: float):
    old_total = (item.cost_price or 0.0) * item.quantity
    new_total = new_unit_cost * new_qty
    total_qty = item.quantity + new_qty
    if total_qty > 0:
        item.cost_price = (old_total + new_total) / total_qty


def _xlsx_response(wb: openpyxl.Workbook, filename: str):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def _style_header(ws, row: int):
    fill = PatternFill("solid", fgColor="1a1a1e")
    font = Font(bold=True, color="C8A060", size=10)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")


# ─────────────────────────────────────────────────────────────────────────────
# ОСТАТКИ
# ─────────────────────────────────────────────────────────────────────────────

@warehouse.route("/")
@login_required
def index():
    q = request.args.get("q", "").strip()
    cat_id = request.args.get("cat", type=int)
    low_stock = request.args.get("low", type=int)

    query = Item.query
    if q:
        query = query.filter(
            db.or_(Item.name.ilike(f"%{q}%"), Item.article.ilike(f"%{q}%"))
        )
    if cat_id:
        query = query.filter(Item.category_id == cat_id)
    if low_stock:
        query = query.filter(Item.quantity <= Item.min_quantity)

    items = query.order_by(Item.name).all()
    categories = Category.query.order_by(Category.name).all()

    all_items = Item.query.all()
    total_items = len(all_items)
    low_items = sum(1 for i in all_items if i.is_low_stock)
    total_cost_sum = sum(i.total_cost for i in all_items)
    reserved_count = sum(1 for i in all_items if i.reserved_qty > 0)

    return render_template(
        "warehouse/index.html",
        items=items,
        categories=categories,
        q=q,
        cat_id=cat_id,
        low_stock=low_stock,
        total_items=total_items,
        low_items=low_items,
        total_cost_sum=total_cost_sum,
        reserved_count=reserved_count,
        active_tab="stock",
    )


@warehouse.route("/export/stock")
@login_required
def export_stock():
    """Экспорт остатков в Excel."""
    items = Item.query.order_by(Item.name).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Остатки"

    ws.merge_cells("A1:L1")
    ws["A1"] = f"Остатки склада — {date.today().strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(bold=True, size=13, color="C8A060")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 24

    ws.append(["Наименование", "Артикул", "Категория", "Ед.",
               "Остаток", "Резерв", "Ожидание", "Доступно",
               "Мин. остаток", "Себестоимость", "Цена", "Место"])
    _style_header(ws, 2)

    for item in items:
        ws.append([
            item.name,
            item.article or "",
            item.category.name if item.category else "",
            item.unit,
            item.quantity,
            item.reserved_qty,
            item.incoming_qty,
            item.available_qty,
            item.min_quantity,
            item.cost_price or 0,
            item.sale_price or 0,
            item.location or "",
        ])
        if item.is_low_stock:
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor="2a1010")

    for i, w in enumerate([40,14,18,6,10,10,10,10,12,14,14,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=3, min_col=10, max_col=11):
        for cell in row:
            cell.number_format = '#,##0.00 "₽"'
    ws.freeze_panes = "A3"

    return _xlsx_response(wb, f"stock_{date.today().isoformat()}.xlsx")


# ─── Управление позициями ────────────────────────────────────────────────────

@warehouse.route("/items/new", methods=["GET", "POST"])
@login_required
def create_item():
    categories = Category.query.order_by(Category.name).all()
    if request.method == "POST":
        item = Item(
            name=request.form["name"].strip(),
            article=request.form.get("article", "").strip() or None,
            category_id=request.form.get("category_id", type=int) or None,
            unit=request.form.get("unit", "шт"),
            quantity=float(request.form.get("quantity", 0)),
            min_quantity=float(request.form.get("min_quantity", 0)),
            cost_price=float(request.form.get("cost_price", 0) or 0),
            sale_price=float(request.form.get("sale_price", 0) or 0),
            location=request.form.get("location", "").strip() or None,
            notes=request.form.get("notes", "").strip() or None,
        )
        db.session.add(item)
        db.session.commit()
        flash("Позиция добавлена", "success")
        return redirect(url_for("warehouse.index"))
    return render_template("warehouse/item_form.html", item=None, categories=categories)


@warehouse.route("/items/<int:item_id>/edit", methods=["GET", "POST"])
@login_required
def edit_item(item_id):
    item = Item.query.get_or_404(item_id)
    categories = Category.query.order_by(Category.name).all()
    if request.method == "POST":
        item.name = request.form["name"].strip()
        item.article = request.form.get("article", "").strip() or None
        item.category_id = request.form.get("category_id", type=int) or None
        item.unit = request.form.get("unit", "шт")
        item.min_quantity = float(request.form.get("min_quantity", 0))
        item.cost_price = float(request.form.get("cost_price", 0) or 0)
        item.sale_price = float(request.form.get("sale_price", 0) or 0)
        item.location = request.form.get("location", "").strip() or None
        item.notes = request.form.get("notes", "").strip() or None
        db.session.commit()
        flash("Позиция обновлена", "success")
        return redirect(url_for("warehouse.index"))
    return render_template("warehouse/item_form.html", item=item, categories=categories)


@warehouse.route("/items/<int:item_id>/delete", methods=["POST"])
@login_required
def delete_item(item_id):
    item = Item.query.get_or_404(item_id)
    db.session.delete(item)
    db.session.commit()
    flash("Позиция удалена", "success")
    return redirect(url_for("warehouse.index"))


# ─── Категории ───────────────────────────────────────────────────────────────

@warehouse.route("/categories/add", methods=["POST"])
@login_required
def add_category():
    name = request.form.get("name", "").strip()
    if name:
        if not Category.query.filter_by(name=name).first():
            db.session.add(Category(name=name))
            db.session.commit()
            flash("Категория добавлена", "success")
        else:
            flash("Категория уже существует", "warning")
    return redirect(url_for("warehouse.index"))


# ─────────────────────────────────────────────────────────────────────────────
# ПРИХОДОВАНИЯ
# ─────────────────────────────────────────────────────────────────────────────

@warehouse.route("/receipts")
@login_required
def receipts():
    status_filter = request.args.get("status", "")
    query = Receipt.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    docs = query.order_by(Receipt.created_at.desc()).all()
    return render_template("warehouse/receipts.html", docs=docs,
                           status_filter=status_filter, active_tab="receipts")


@warehouse.route("/receipts/new", methods=["GET", "POST"])
@login_required
def create_receipt():
    items = Item.query.order_by(Item.name).all()
    if request.method == "POST":
        number = _next_number("ПРХ", Receipt)
        receipt_date = None
        raw = request.form.get("receipt_date", "").strip()
        if raw:
            try:
                receipt_date = datetime.strptime(raw, "%Y-%m-%d")
            except ValueError:
                pass

        doc = Receipt(
            number=number,
            supplier=request.form.get("supplier", "").strip() or None,
            notes=request.form.get("notes", "").strip() or None,
            receipt_date=receipt_date,
            status="draft",
            created_by_id=current_user.id,
        )
        db.session.add(doc)
        db.session.flush()

        for iid, qty_s, cost_s in zip(
            request.form.getlist("item_id[]"),
            request.form.getlist("quantity[]"),
            request.form.getlist("unit_cost[]"),
        ):
            if not iid or not qty_s:
                continue
            db.session.add(ReceiptItem(
                receipt_id=doc.id,
                item_id=int(iid),
                quantity=float(qty_s),
                unit_cost=float(cost_s or 0),
            ))

        db.session.commit()
        flash(f"Приходование {number} создано (черновик)", "success")
        return redirect(url_for("warehouse.receipts"))

    return render_template("warehouse/receipt_form.html", doc=None, items=items,
                           today=date.today().isoformat(), active_tab="receipts")


@warehouse.route("/receipts/<int:doc_id>/confirm", methods=["POST"])
@login_required
def confirm_receipt(doc_id):
    doc = Receipt.query.get_or_404(doc_id)
    if doc.status != "draft":
        flash("Документ уже проведён или отменён", "warning")
        return redirect(url_for("warehouse.receipts"))
    for ri in doc.items:
        _update_avg_cost(ri.item, ri.quantity, ri.unit_cost or 0)
        ri.item.quantity += ri.quantity
        ri.item.incoming_qty = max(ri.item.incoming_qty - ri.quantity, 0)
        db.session.add(StockMovement(
            item_id=ri.item.id, move_type="receipt",
            quantity=ri.quantity, unit_cost=ri.unit_cost,
            document_ref=doc.number,
            notes=f"Приход от {doc.supplier or '—'}",
            created_by_id=current_user.id,
        ))
    doc.status = "confirmed"
    db.session.commit()
    flash(f"Приходование {doc.number} проведено", "success")
    return redirect(url_for("warehouse.receipts"))


@warehouse.route("/receipts/<int:doc_id>/cancel", methods=["POST"])
@login_required
def cancel_receipt(doc_id):
    doc = Receipt.query.get_or_404(doc_id)
    if doc.status == "confirmed":
        flash("Нельзя отменить проведённый документ", "error")
        return redirect(url_for("warehouse.receipts"))
    doc.status = "cancelled"
    db.session.commit()
    flash(f"Приходование {doc.number} отменено", "success")
    return redirect(url_for("warehouse.receipts"))


@warehouse.route("/receipts/<int:doc_id>/export")
@login_required
def export_receipt(doc_id):
    doc = Receipt.query.get_or_404(doc_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Приходование"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Приходование {doc.number}"
    ws["A1"].font = Font(bold=True, size=14, color="C8A060")
    ws["A1"].alignment = Alignment(horizontal="center")

    for label, val, col in [("Поставщик:", doc.supplier or "—", "B2"),
                              ("Дата:", (doc.receipt_date or doc.created_at).strftime("%d.%m.%Y"), "E2"),
                              ("Статус:", doc.status_label, "B3")]:
        pass
    ws["A2"] = "Поставщик:"; ws["A2"].font = Font(bold=True)
    ws["B2"] = doc.supplier or "—"
    ws["D2"] = "Дата:"; ws["D2"].font = Font(bold=True)
    ws["E2"] = (doc.receipt_date or doc.created_at).strftime("%d.%m.%Y")
    ws["A3"] = "Статус:"; ws["A3"].font = Font(bold=True)
    ws["B3"] = doc.status_label

    ws.append([])
    ws.append(["Наименование", "Артикул", "Ед.", "Количество", "Себестоимость", "Сумма"])
    _style_header(ws, 5)

    total = 0
    for ri in doc.items:
        summa = ri.quantity * (ri.unit_cost or 0)
        total += summa
        ws.append([
            ri.item.name if ri.item else "—",
            ri.item.article or "" if ri.item else "",
            ri.item.unit if ri.item else "",
            ri.quantity, ri.unit_cost or 0, summa,
        ])
        ws.cell(ws.max_row, 5).number_format = '#,##0.00 "₽"'
        ws.cell(ws.max_row, 6).number_format = '#,##0.00 "₽"'

    tr = ws.max_row + 2
    ws.cell(tr, 5, "Итого:").font = Font(bold=True)
    ws.cell(tr, 6, total).font = Font(bold=True)
    ws.cell(tr, 6).number_format = '#,##0.00 "₽"'

    for i, w in enumerate([40,14,6,12,16,16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"
    return _xlsx_response(wb, f"receipt_{doc.number}.xlsx")


@warehouse.route("/receipts/<int:doc_id>/print")
@login_required
def print_receipt(doc_id):
    doc = Receipt.query.get_or_404(doc_id)
    return render_template("warehouse/print/receipt.html", doc=doc)


# ─────────────────────────────────────────────────────────────────────────────
# СПИСАНИЯ
# ─────────────────────────────────────────────────────────────────────────────

@warehouse.route("/write-offs")
@login_required
def write_offs():
    status_filter = request.args.get("status", "")
    query = WriteOff.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    docs = query.order_by(WriteOff.created_at.desc()).all()
    return render_template("warehouse/write_offs.html", docs=docs,
                           status_filter=status_filter, active_tab="write_offs")


@warehouse.route("/write-offs/new", methods=["GET", "POST"])
@login_required
def create_write_off():
    items = Item.query.order_by(Item.name).all()
    if request.method == "POST":
        number = _next_number("СПС", WriteOff)
        doc = WriteOff(
            number=number,
            reason=request.form.get("reason", "other"),
            notes=request.form.get("notes", "").strip() or None,
            status="draft",
            created_by_id=current_user.id,
        )
        db.session.add(doc)
        db.session.flush()
        for iid, qty_s in zip(request.form.getlist("item_id[]"),
                               request.form.getlist("quantity[]")):
            if not iid or not qty_s:
                continue
            db.session.add(WriteOffItem(
                write_off_id=doc.id, item_id=int(iid), quantity=float(qty_s)
            ))
        db.session.commit()
        flash(f"Списание {number} создано (черновик)", "success")
        return redirect(url_for("warehouse.write_offs"))
    return render_template("warehouse/write_off_form.html", doc=None, items=items,
                           active_tab="write_offs")


@warehouse.route("/write-offs/<int:doc_id>/confirm", methods=["POST"])
@login_required
def confirm_write_off(doc_id):
    doc = WriteOff.query.get_or_404(doc_id)
    if doc.status != "draft":
        flash("Документ уже проведён или отменён", "warning")
        return redirect(url_for("warehouse.write_offs"))
    errors = [
        f"«{wi.item.name}»: недостаточно (есть {wi.item.quantity}, нужно {wi.quantity})"
        for wi in doc.items if wi.item.quantity < wi.quantity
    ]
    if errors:
        for e in errors:
            flash(e, "error")
        return redirect(url_for("warehouse.write_offs"))
    for wi in doc.items:
        wi.item.quantity -= wi.quantity
        db.session.add(StockMovement(
            item_id=wi.item.id, move_type="write_off",
            quantity=-wi.quantity, unit_cost=wi.item.cost_price,
            document_ref=doc.number, notes=f"Списание: {doc.reason_label}",
            created_by_id=current_user.id,
        ))
    doc.status = "confirmed"
    db.session.commit()
    flash(f"Списание {doc.number} проведено", "success")
    return redirect(url_for("warehouse.write_offs"))


@warehouse.route("/write-offs/<int:doc_id>/cancel", methods=["POST"])
@login_required
def cancel_write_off(doc_id):
    doc = WriteOff.query.get_or_404(doc_id)
    if doc.status == "confirmed":
        flash("Нельзя отменить проведённый документ", "error")
        return redirect(url_for("warehouse.write_offs"))
    doc.status = "cancelled"
    db.session.commit()
    flash(f"Списание {doc.number} отменено", "success")
    return redirect(url_for("warehouse.write_offs"))


@warehouse.route("/write-offs/<int:doc_id>/export")
@login_required
def export_write_off(doc_id):
    doc = WriteOff.query.get_or_404(doc_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Списание"

    ws.merge_cells("A1:E1")
    ws["A1"] = f"Акт списания {doc.number}"
    ws["A1"].font = Font(bold=True, size=14, color="C8A060")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "Причина:"; ws["A2"].font = Font(bold=True); ws["B2"] = doc.reason_label
    ws["D2"] = "Дата:"; ws["D2"].font = Font(bold=True)
    ws["E2"] = doc.created_at.strftime("%d.%m.%Y")

    ws.append([])
    ws.append(["Наименование", "Артикул", "Ед.", "Количество", "Сумма списания"])
    _style_header(ws, 4)

    total = 0
    for wi in doc.items:
        summa = wi.quantity * (wi.item.cost_price or 0) if wi.item else 0
        total += summa
        ws.append([
            wi.item.name if wi.item else "—",
            wi.item.article or "" if wi.item else "",
            wi.item.unit if wi.item else "",
            wi.quantity, summa,
        ])
        ws.cell(ws.max_row, 5).number_format = '#,##0.00 "₽"'

    tr = ws.max_row + 2
    ws.cell(tr, 4, "Итого:").font = Font(bold=True)
    ws.cell(tr, 5, total).font = Font(bold=True)
    ws.cell(tr, 5).number_format = '#,##0.00 "₽"'

    for i, w in enumerate([40,14,6,12,18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    return _xlsx_response(wb, f"write_off_{doc.number}.xlsx")


@warehouse.route("/write-offs/<int:doc_id>/print")
@login_required
def print_write_off(doc_id):
    doc = WriteOff.query.get_or_404(doc_id)
    return render_template("warehouse/print/write_off.html", doc=doc)


# ─────────────────────────────────────────────────────────────────────────────
# ПЕРЕМЕЩЕНИЯ
# ─────────────────────────────────────────────────────────────────────────────

@warehouse.route("/movements")
@login_required
def movements():
    q = request.args.get("q", "").strip()
    move_type = request.args.get("type", "")
    page = request.args.get("page", 1, type=int)

    query = StockMovement.query
    if move_type:
        query = query.filter_by(move_type=move_type)
    if q:
        query = query.join(Item).filter(Item.name.ilike(f"%{q}%"))

    pagination = query.order_by(StockMovement.created_at.desc()).paginate(
        page=page, per_page=PER_PAGE, error_out=False
    )
    items = Item.query.order_by(Item.name).all()

    return render_template(
        "warehouse/movements.html",
        docs=pagination.items,
        pagination=pagination,
        items=items,
        q=q,
        move_type=move_type,
        move_types=StockMovement.TYPES,
        active_tab="movements",
    )


@warehouse.route("/movements/new", methods=["POST"])
@login_required
def create_movement():
    item_id = request.form.get("item_id", type=int)
    qty = float(request.form.get("quantity", 0))
    from_loc = request.form.get("from_location", "").strip() or None
    to_loc = request.form.get("to_location", "").strip() or None

    item = Item.query.get_or_404(item_id)
    if item.quantity < qty:
        flash(f"Недостаточно остатка (есть {item.quantity} {item.unit})", "error")
        return redirect(url_for("warehouse.movements"))

    item.location = to_loc or item.location
    db.session.add(StockMovement(
        item_id=item.id, move_type="transfer", quantity=qty,
        from_location=from_loc, to_location=to_loc,
        notes=request.form.get("notes", "").strip() or None,
        created_by_id=current_user.id,
    ))
    db.session.commit()
    flash(f"Перемещение «{item.name}» зарегистрировано", "success")
    return redirect(url_for("warehouse.movements"))


@warehouse.route("/movements/export")
@login_required
def export_movements():
    move_type = request.args.get("type", "")
    query = StockMovement.query
    if move_type:
        query = query.filter_by(move_type=move_type)
    docs = query.order_by(StockMovement.created_at.desc()).limit(1000).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Движения"

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Журнал движений — {date.today().strftime('%d.%m.%Y')}"
    ws["A1"].font = Font(bold=True, size=13, color="C8A060")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append(["Дата", "Тип", "Позиция", "Кол-во", "Откуда", "Куда", "Документ", "Автор"])
    _style_header(ws, 2)

    for m in docs:
        ws.append([
            m.created_at.strftime("%d.%m.%Y %H:%M"),
            m.type_label,
            m.item.name if m.item else "—",
            m.quantity,
            m.from_location or "",
            m.to_location or "",
            m.document_ref or "",
            m.created_by.full_name if m.created_by else "—",
        ])

    for i, w in enumerate([18,16,36,10,18,18,16,20], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"
    return _xlsx_response(wb, f"movements_{date.today().isoformat()}.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# ИНВЕНТАРИЗАЦИИ
# ─────────────────────────────────────────────────────────────────────────────

@warehouse.route("/inventory")
@login_required
def inventory():
    docs = InventoryCheck.query.order_by(InventoryCheck.created_at.desc()).all()
    return render_template("warehouse/inventory.html", docs=docs, active_tab="inventory")


@warehouse.route("/inventory/new", methods=["POST"])
@login_required
def create_inventory():
    number = _next_number("ИНВ", InventoryCheck)
    check = InventoryCheck(
        number=number, status="in_progress",
        notes=request.form.get("notes", "").strip() or None,
        created_by_id=current_user.id,
    )
    db.session.add(check)
    db.session.flush()
    for item in Item.query.all():
        db.session.add(InventoryCheckItem(
            check_id=check.id, item_id=item.id,
            expected_qty=item.quantity, actual_qty=None,
        ))
    db.session.commit()
    flash(f"Инвентаризация {number} начата", "success")
    return redirect(url_for("warehouse.inventory_detail", check_id=check.id))


@warehouse.route("/inventory/<int:check_id>")
@login_required
def inventory_detail(check_id):
    check = InventoryCheck.query.get_or_404(check_id)
    return render_template("warehouse/inventory_detail.html", check=check, active_tab="inventory")


@warehouse.route("/inventory/<int:check_id>/save", methods=["POST"])
@login_required
def save_inventory(check_id):
    check = InventoryCheck.query.get_or_404(check_id)
    if check.status != "in_progress":
        flash("Инвентаризация уже завершена", "warning")
        return redirect(url_for("warehouse.inventory_detail", check_id=check_id))
    for line in check.lines:
        val = request.form.get(f"actual_{line.id}", "").strip()
        line.actual_qty = float(val) if val != "" else None
    db.session.commit()
    flash("Данные сохранены", "success")
    return redirect(url_for("warehouse.inventory_detail", check_id=check_id))


@warehouse.route("/inventory/<int:check_id>/finish", methods=["POST"])
@login_required
def finish_inventory(check_id):
    check = InventoryCheck.query.get_or_404(check_id)
    if check.status != "in_progress":
        flash("Инвентаризация уже завершена или отменена", "warning")
        return redirect(url_for("warehouse.inventory_detail", check_id=check_id))
    for line in check.lines:
        val = request.form.get(f"actual_{line.id}", "").strip()
        line.actual_qty = float(val) if val != "" else None
    for line in check.lines:
        if line.actual_qty is None:
            continue
        diff = line.actual_qty - line.expected_qty
        if diff == 0:
            continue
        item = line.item
        item.quantity = max(item.quantity + diff, 0)
        db.session.add(StockMovement(
            item_id=item.id, move_type="adjustment", quantity=diff,
            unit_cost=item.cost_price, document_ref=check.number,
            notes=f"Корректировка по инвентаризации {check.number}",
            created_by_id=current_user.id,
        ))
    check.status = "done"
    check.finished_at = datetime.utcnow()
    db.session.commit()
    flash(f"Инвентаризация {check.number} завершена", "success")
    return redirect(url_for("warehouse.inventory"))


@warehouse.route("/inventory/<int:check_id>/cancel", methods=["POST"])
@login_required
def cancel_inventory(check_id):
    check = InventoryCheck.query.get_or_404(check_id)
    if check.status == "done":
        flash("Нельзя отменить завершённую инвентаризацию", "error")
        return redirect(url_for("warehouse.inventory"))
    check.status = "cancelled"
    db.session.commit()
    flash(f"Инвентаризация {check.number} отменена", "success")
    return redirect(url_for("warehouse.inventory"))


@warehouse.route("/inventory/<int:check_id>/export")
@login_required
def export_inventory(check_id):
    check = InventoryCheck.query.get_or_404(check_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Инвентаризация"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Инвентаризация {check.number}"
    ws["A1"].font = Font(bold=True, size=14, color="C8A060")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "Дата:"; ws["A2"].font = Font(bold=True)
    ws["B2"] = check.created_at.strftime("%d.%m.%Y")
    ws["D2"] = "Статус:"; ws["D2"].font = Font(bold=True)
    ws["E2"] = check.status_label

    ws.append([])
    ws.append(["Наименование", "Артикул", "Категория", "Ед.", "По учёту", "Факт", "Расхождение"])
    _style_header(ws, 4)

    for line in check.lines:
        item = line.item
        diff = line.diff
        ws.append([
            item.name if item else "—",
            item.article or "" if item else "",
            item.category.name if item and item.category else "",
            item.unit if item else "",
            line.expected_qty,
            line.actual_qty if line.actual_qty is not None else "",
            diff if diff is not None else "",
        ])
        if diff is not None and diff != 0:
            clr = "2a1010" if diff < 0 else "0d2a0d"
            for cell in ws[ws.max_row]:
                cell.fill = PatternFill("solid", fgColor=clr)

    for i, w in enumerate([40,14,18,6,12,12,14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A5"
    return _xlsx_response(wb, f"inventory_{check.number}.xlsx")


@warehouse.route("/inventory/<int:check_id>/print")
@login_required
def print_inventory(check_id):
    check = InventoryCheck.query.get_or_404(check_id)
    return render_template("warehouse/print/inventory.html", check=check)


# ─────────────────────────────────────────────────────────────────────────────
# ЗАКАЗЫ
# ─────────────────────────────────────────────────────────────────────────────

@warehouse.route("/orders")
@login_required
def orders():
    status_filter = request.args.get("status", "")
    query = Order.query
    if status_filter:
        query = query.filter_by(status=status_filter)
    all_orders = query.order_by(Order.created_at.desc()).all()
    return render_template("warehouse/orders.html", orders=all_orders,
                           status_filter=status_filter, active_tab="orders")


@warehouse.route("/orders/new", methods=["GET", "POST"])
@login_required
def create_order():
    items = Item.query.order_by(Item.name).all()
    engineers = User.query.filter(User.role.in_(["engineer", "director"])).all()
    if request.method == "POST":
        number = _next_number("ЗКЗ", Order)
        order = Order(
            number=number,
            object_name=request.form["object_name"].strip(),
            notes=request.form.get("notes", "").strip() or None,
            status="new",
            created_by_id=current_user.id,
            assigned_to_id=request.form.get("assigned_to_id", type=int) or None,
        )
        db.session.add(order)
        db.session.flush()
        for iid, qty_s in zip(request.form.getlist("item_id[]"),
                               request.form.getlist("quantity[]")):
            if not iid or not qty_s:
                continue
            qty = float(qty_s)
            db.session.add(OrderItem(order_id=order.id, item_id=int(iid), quantity=qty))
            item = Item.query.get(int(iid))
            if item:
                item.reserved_qty = (item.reserved_qty or 0) + qty
        db.session.commit()
        flash(f"Заказ {number} создан", "success")
        return redirect(url_for("warehouse.orders"))
    return render_template("warehouse/order_form.html", items=items,
                           engineers=engineers, active_tab="orders")


@warehouse.route("/orders/<int:order_id>/status", methods=["POST"])
@login_required
def update_order_status(order_id):
    order = Order.query.get_or_404(order_id)
    new_status = request.form.get("status")
    old_status = order.status

    if new_status == "issued" and old_status != "issued":
        for oi in order.items:
            item = oi.item
            qty = oi.issued_quantity or oi.quantity
            item.quantity = max(item.quantity - qty, 0)
            item.reserved_qty = max(item.reserved_qty - oi.quantity, 0)
            db.session.add(StockMovement(
                item_id=item.id, move_type="order_out", quantity=-qty,
                unit_cost=item.cost_price, order_id=order.id,
                document_ref=order.number,
                notes=f"Выдача по заказу {order.number}",
                created_by_id=current_user.id,
            ))
        order.issued_at = datetime.utcnow()
    elif new_status == "cancelled" and old_status not in ("issued", "cancelled"):
        for oi in order.items:
            if oi.item:
                oi.item.reserved_qty = max(oi.item.reserved_qty - oi.quantity, 0)

    if new_status in Order.STATUS_CHOICES:
        order.status = new_status
    db.session.commit()
    flash(f"Статус заказа {order.number} изменён", "success")
    return redirect(url_for("warehouse.orders"))


@warehouse.route("/orders/<int:order_id>/export")
@login_required
def export_order(order_id):
    order = Order.query.get_or_404(order_id)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заказ"

    ws.merge_cells("A1:F1")
    ws["A1"] = f"Заказ {order.number}"
    ws["A1"].font = Font(bold=True, size=14, color="C8A060")
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"] = "Объект:"; ws["A2"].font = Font(bold=True); ws["B2"] = order.object_name
    ws["D2"] = "Инженер:"; ws["D2"].font = Font(bold=True)
    ws["E2"] = order.assignee.full_name if order.assignee else "—"
    ws["A3"] = "Статус:"; ws["A3"].font = Font(bold=True); ws["B3"] = order.status_label
    ws["D3"] = "Дата:"; ws["D3"].font = Font(bold=True)
    ws["E3"] = order.created_at.strftime("%d.%m.%Y")

    ws.append([])
    ws.append(["Наименование", "Артикул", "Ед.", "Заказано", "Выдано", "Остаток"])
    _style_header(ws, 5)

    for oi in order.items:
        item = oi.item
        issued = oi.issued_quantity or 0
        ws.append([
            item.name if item else "—",
            item.article or "" if item else "",
            item.unit if item else "",
            oi.quantity, issued, oi.quantity - issued,
        ])

    for i, w in enumerate([40,14,6,12,12,12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A6"
    return _xlsx_response(wb, f"order_{order.number}.xlsx")


@warehouse.route("/orders/<int:order_id>/print")
@login_required
def print_order(order_id):
    order = Order.query.get_or_404(order_id)
    return render_template("warehouse/print/order.html", order=order)